from decimal import Decimal
import datetime
import io
import re
import shutil
import uuid
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from starlette.datastructures import UploadFile as StarletteUploadFile
from sqlalchemy import func, or_, and_
from sqlalchemy.orm import Session, joinedload

import models
import schemas
from api.auth import get_current_user
from database import get_db
from models import ComplaintResolutionStatus
from utils.audit import log_audit
from utils.auth import hash_password
from utils.cache import (
    cache_get, cache_set, cache_delete, cache_delete_pattern,
    CACHE_KEY_ADMIN_STATS, CACHE_TTL_ADMIN_STATS,
    CACHE_KEY_CATEGORIES, CACHE_KEY_PRODUCTS_PREFIX,
)
from utils.freshness import ensure_delivered_notification
from utils.inventory import apply_stock_delta, set_product_quantity, derive_stock_status, record_stock_transaction

router = APIRouter(tags=["Admin"])
PRODUCT_UPLOAD_DIR = Path(__file__).resolve().parent.parent / "uploads" / "products"
PAYMENT_QR_UPLOAD_DIR = Path(__file__).resolve().parent.parent / "uploads" / "payment_qr"
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}

STATUS_ALIASES = {
    "Cho xac nhan": models.OrderStatus.PENDING,
    "Chờ xác nhận": models.OrderStatus.PENDING,
    "pending": models.OrderStatus.PENDING,
    "Da xac nhan": models.OrderStatus.CONFIRMED,
    "Đã xác nhận": models.OrderStatus.CONFIRMED,
    "confirmed": models.OrderStatus.CONFIRMED,
    "Dang giao": models.OrderStatus.SHIPPED,
    "Đang giao": models.OrderStatus.SHIPPED,
    "shipped": models.OrderStatus.SHIPPED,
    "Da giao": models.OrderStatus.DELIVERED,
    "Đã giao": models.OrderStatus.DELIVERED,
    "delivered": models.OrderStatus.DELIVERED,
    "Da huy": models.OrderStatus.CANCELLED,
    "Đã hủy": models.OrderStatus.CANCELLED,
    "cancelled": models.OrderStatus.CANCELLED,
    "returned": models.OrderStatus.RETURNED,
}


class AdminProductCreate(BaseModel):
    name: str
    description: Optional[str] = None
    price: Decimal
    category_id: Optional[int] = None
    category: Optional[str] = None
    image_url: Optional[str] = None
    img: Optional[str] = None
    stock: Optional[int] = Field(default=None, ge=0)
    quantity: Optional[int] = Field(default=None, ge=0)
    unit: Optional[str] = "kg"
    stock_status: Optional[str] = None
    rating: Optional[float] = Field(default=5.0, ge=0, le=5)
    low_stock_threshold: Optional[int] = Field(default=5, ge=0)


class AdminProductUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    price: Optional[Decimal] = None
    category_id: Optional[int] = None
    category: Optional[str] = None
    image_url: Optional[str] = None
    img: Optional[str] = None
    stock: Optional[int] = Field(default=None, ge=0)
    quantity: Optional[int] = Field(default=None, ge=0)
    unit: Optional[str] = None
    stock_status: Optional[str] = None
    rating: Optional[float] = Field(default=None, ge=0, le=5)
    low_stock_threshold: Optional[int] = Field(default=None, ge=0)


def get_current_admin_user(
    current_user: models.User = Depends(get_current_user),
) -> models.User:
    if not current_user.is_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Ban khong co quyen truy cap tai nguyen Admin",
        )
    return current_user


def get_current_staff_or_admin_user(
    current_user: models.User = Depends(get_current_user),
) -> models.User:
    """Cho phép cả Admin và Nhân viên (Staff) truy cập."""
    if not (current_user.is_admin or current_user.is_staff):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Ban khong co quyen thuc hien thao tac nay",
        )
    return current_user


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return slug or "item"


def normalize_order_status(value: str) -> models.OrderStatus:
    if value in STATUS_ALIASES:
        return STATUS_ALIASES[value]
    lowered = value.lower()
    if lowered in STATUS_ALIASES:
        return STATUS_ALIASES[lowered]
    valid = ", ".join(status.value for status in models.OrderStatus)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=f"Trang thai khong hop le. Hay chon mot trong: {valid}",
    )


def apply_order_status_transition(order: models.Order, new_status: models.OrderStatus) -> None:
    """Keep status timestamps consistent for both single and bulk admin updates."""
    now = datetime.datetime.utcnow()
    order.status = new_status
    order.updated_at = now

    if new_status == models.OrderStatus.SHIPPED:
        order.shipped_at = order.shipped_at or now
        order.delivered_at = None
        return

    if new_status == models.OrderStatus.DELIVERED:
        order.shipped_at = order.shipped_at or now
        order.delivered_at = now
        return

    if new_status in {models.OrderStatus.PENDING, models.OrderStatus.CONFIRMED}:
        order.shipped_at = None
        order.delivered_at = None
        return

    if new_status == models.OrderStatus.CANCELLED:
        order.delivered_at = None


def build_admin_orders_query(db: Session):
    return (
        db.query(models.Order)
        .join(models.User, models.Order.user_id == models.User.id)
        .options(
            joinedload(models.Order.items).joinedload(models.OrderItem.product),
            joinedload(models.Order.owner),
        )
    )


def get_unread_feedback_filter():
    return or_(
        models.ScanFeedbackEvent.is_read == False,
        models.ScanFeedbackEvent.is_read.is_(None),
    )


def serialize_user_brief(user: Optional[models.User]):
    if user is None:
        return None

    return {
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "full_name": user.full_name,
        "is_admin": bool(user.is_admin),
        "is_staff": bool(user.is_staff),
        "role": user.role.value if hasattr(user.role, "value") else user.role,
    }


def serialize_feedback_event(event: models.ScanFeedbackEvent):
    return {
        "id": event.id,
        "scan_id": event.scan_id,
        "user_id": event.user_id,
        "source": event.source,
        "image_url": event.image_url,
        "model_version": event.model_version,
        "predicted_label": event.predicted_label,
        "predicted_status": event.predicted_status,
        "predicted_confidence": float(event.predicted_confidence or 0),
        "is_correct": bool(event.is_correct),
        "corrected_label": event.corrected_label,
        "corrected_status": event.corrected_status,
        "notes": event.notes,
        "extra_metadata": event.extra_metadata,
        "is_read": bool(event.is_read),
        "read_at": event.read_at,
        "created_at": event.created_at,
        "user": serialize_user_brief(event.user),
        "scan": (
            {
                "id": event.scan.id,
                "image_url": event.scan.image_url,
                "created_at": event.scan.created_at,
            }
            if event.scan
            else None
        ),
    }


def serialize_payment_qr_setting(setting: models.PaymentQRCodeSetting):
    return {
        "id": setting.id,
        "provider_name": setting.provider_name,
        "image_url": setting.image_url,
        "updated_by_user_id": setting.updated_by_user_id,
        "updated_by": serialize_user_brief(setting.updated_by),
        "created_at": setting.created_at,
        "updated_at": setting.updated_at,
    }


def build_page_meta(page: int, limit: int, total: int) -> tuple[int, int, bool]:
    total_pages = max(1, (total + limit - 1) // limit)
    safe_page = min(page, total_pages) if total else 1
    has_next = safe_page < total_pages
    return safe_page, total_pages, has_next


def apply_admin_order_filters(
    query,
    *,
    search: Optional[str] = None,
    order_status: Optional[str] = None,
    payment_method: Optional[str] = None,
    date_from: Optional[datetime.date] = None,
    date_to: Optional[datetime.date] = None,
):
    if order_status and order_status != "all":
        query = query.filter(models.Order.status == normalize_order_status(order_status))

    if payment_method and payment_method != "all":
        query = query.filter(func.lower(func.coalesce(models.Order.payment_method, "")) == payment_method.strip().lower())

    if search:
        keyword = f"%{search.strip().lower()}%"
        query = query.filter(
            or_(
                func.lower(func.coalesce(models.Order.order_number, "")).like(keyword),
                func.lower(func.coalesce(models.User.full_name, "")).like(keyword),
                func.lower(func.coalesce(models.User.username, "")).like(keyword),
            )
        )

    if date_from:
        start_at = datetime.datetime.combine(date_from, datetime.time.min)
        query = query.filter(models.Order.created_at >= start_at)

    if date_to:
        end_at = datetime.datetime.combine(date_to + datetime.timedelta(days=1), datetime.time.min)
        query = query.filter(models.Order.created_at < end_at)

    return query


def apply_admin_user_filters(
    query,
    *,
    search: Optional[str] = None,
    role: Optional[str] = None,
):
    normalized_role = (role or "").strip().lower()
    if normalized_role and normalized_role != "all":
        query = query.filter(models.User.role == normalized_role)

    if search:
        keyword = f"%{search.strip().lower()}%"
        query = query.filter(
            or_(
                func.lower(func.coalesce(models.User.username, "")).like(keyword),
                func.lower(func.coalesce(models.User.email, "")).like(keyword),
                func.lower(func.coalesce(models.User.full_name, "")).like(keyword),
            )
        )

    return query


def apply_stock_transaction_filters(
    query,
    *,
    search: Optional[str] = None,
    transaction_type: Optional[str] = None,
):
    normalized_type = (transaction_type or "").strip().lower()
    if normalized_type and normalized_type != "all":
        query = query.filter(models.StockTransaction.type == normalized_type)

    if search:
        keyword = f"%{search.strip().lower()}%"
        query = query.filter(
            or_(
                func.lower(func.coalesce(models.Product.name, "")).like(keyword),
                func.lower(func.coalesce(models.StockTransaction.note, "")).like(keyword),
                func.lower(func.coalesce(models.User.username, "")).like(keyword),
                func.lower(func.coalesce(models.User.full_name, "")).like(keyword),
            )
        )

    return query


def require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Tinh nang Excel chua san sang tren server. Vui long cai dat openpyxl.",
        ) from exc
    return Workbook, load_workbook


def normalize_excel_header(value: object) -> str:
    raw = str(value or "").strip().lower()
    normalized = re.sub(r"[^a-z0-9]+", "_", raw).strip("_")
    return normalized


def normalize_excel_transaction_type(value: object) -> str:
    normalized = str(value or "").strip().lower()
    aliases = {
        "import": "import",
        "nhap": "import",
        "nhap_kho": "import",
        "nhap_hang": "import",
        "in": "import",
        "export": "export",
        "xuat": "export",
        "xuat_kho": "export",
        "xuat_hang": "export",
        "out": "export",
    }
    if normalized not in aliases:
        raise ValueError("Loai giao dich phai la import hoac export")
    return aliases[normalized]


def parse_excel_transaction_date(value: object) -> Optional[datetime.datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)

    raw = str(value).strip()
    if not raw:
        return None

    for parser in (datetime.datetime.fromisoformat,):
        try:
            return parser(raw)
        except ValueError:
            continue

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            parsed = datetime.datetime.strptime(raw, fmt)
            return parsed
        except ValueError:
            continue

    raise ValueError("Ngay giao dich khong dung dinh dang hop le")


def build_excel_error_detail(prefix: str, errors: List[str]) -> str:
    preview = errors[:8]
    detail = "\n".join(preview)
    if len(errors) > len(preview):
        detail = f"{detail}\n... va {len(errors) - len(preview)} dong loi khac"
    return f"{prefix}\n{detail}"


def ensure_upload_dir(upload_dir: Path):
    upload_dir.mkdir(parents=True, exist_ok=True)


def save_uploaded_image(image_file: UploadFile, *, upload_dir: Path, relative_dir: str) -> str:
    if not image_file.filename:
        raise HTTPException(status_code=400, detail="Image file name is missing")

    extension = Path(image_file.filename).suffix.lower()
    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_IMAGE_EXTENSIONS))
        raise HTTPException(status_code=400, detail=f"Unsupported file type. Allowed: {allowed}")

    ensure_upload_dir(upload_dir)
    file_name = f"{uuid.uuid4().hex}{extension}"
    output_path = upload_dir / file_name

    with output_path.open("wb") as file_stream:
        shutil.copyfileobj(image_file.file, file_stream)

    return f"/uploads/{relative_dir}/{file_name}"


async def parse_product_payload(request: Request):
    content_type = request.headers.get("content-type", "")

    if "multipart/form-data" in content_type:
        form_data = await request.form()
        payload = {
            "name": form_data.get("name"),
            "description": form_data.get("description"),
            "price": form_data.get("price"),
            "category_id": form_data.get("category_id"),
            "category": form_data.get("category"),
            "image_url": form_data.get("image_url"),
            "img": form_data.get("img"),
            "stock": form_data.get("stock"),
            "quantity": form_data.get("quantity"),
            "unit": form_data.get("unit"),
            "stock_status": form_data.get("stock_status"),
            "rating": form_data.get("rating"),
            "low_stock_threshold": form_data.get("low_stock_threshold"),
        }
        image_file = form_data.get("image_file")
        if image_file and not isinstance(image_file, (UploadFile, StarletteUploadFile)):
            # Accept Starlette/FastAPI upload objects; ignore invalid scalar values.
            if not (hasattr(image_file, "filename") and hasattr(image_file, "file")):
                image_file = None
        return payload, image_file

    try:
        payload = await request.json()
    except Exception:
        payload = {}

    return payload, None


def resolve_category_id(
    db: Session,
    category_id: Optional[int],
    category_name: Optional[str],
) -> int:
    if category_id is not None:
        category = db.query(models.Category).filter(models.Category.id == category_id).first()
        if not category:
            raise HTTPException(status_code=404, detail="Danh muc khong ton tai")
        return category.id

    category_value = (category_name or "Khac").strip()
    category = db.query(models.Category).filter(models.Category.name == category_value).first()
    if not category:
        category = models.Category(
            name=category_value,
            slug=slugify(category_value),
            is_active=True,
        )
        db.add(category)
        db.flush()

    return category.id


@router.post("/products/add")
async def add_product(
    request: Request,
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    # [H3] Audit log được ghi sau khi tạo product thành công (xem cuối hàm)
    payload, image_file = await parse_product_payload(request)
    product = AdminProductCreate(**payload)

    # Validate tên sản phẩm không bị trùng (case-insensitive, bỏ khoảng trắng thừa)
    normalized_name = product.name.strip().lower()
    duplicate = db.query(models.Product).filter(
        func.lower(func.trim(models.Product.name)) == normalized_name,
        models.Product.is_active == True,
    ).first()
    if duplicate:
        raise HTTPException(
            status_code=409,
            detail=f"Sản phẩm với tên '{product.name.strip()}' đã tồn tại. Vui lòng chọn tên khác.",
        )

    category_id = resolve_category_id(db, product.category_id, product.category)

    image_url = product.image_url or product.img
    if image_file:
        image_url = save_uploaded_image(
            image_file,
            upload_dir=PRODUCT_UPLOAD_DIR,
            relative_dir="products",
        )

    quantity = product.quantity if product.quantity is not None else (product.stock or 0)
    stock_status = derive_stock_status(quantity)
    low_stock_threshold = product.low_stock_threshold if product.low_stock_threshold is not None else 5

    new_product = models.Product(
        name=product.name.strip(),
        slug=slugify(product.name),
        description=product.description or "",
        price=product.price,
        category_id=category_id,
        image_url=image_url,
        quantity=quantity,
        low_stock_threshold=low_stock_threshold,
        unit=product.unit or "kg",
        stock_status=stock_status,
        rating=product.rating or 5.0,
        is_active=True,
    )

    db.add(new_product)
    db.flush()
    if quantity > 0:
        record_stock_transaction(
            db=db,
            product=new_product,
            user_id=current_user.id,
            transaction_type=models.TransactionType.IMPORT,
            quantity=quantity,
            note=f"Ton kho ban dau khi tao san pham: {new_product.name}",
        )
    # [H3] Audit log
    log_audit(
        db, user_id=current_user.id,
        action="created_product", resource_type="product",
        resource_id=new_product.id,
        details={"name": new_product.name, "price": float(new_product.price), "quantity": quantity},
        request=request,
    )
    db.commit()
    db.refresh(new_product)
    # [M5] Invalidate product and stats caches
    cache_delete_pattern(CACHE_KEY_PRODUCTS_PREFIX + "*")
    cache_delete(CACHE_KEY_ADMIN_STATS)
    cache_delete(CACHE_KEY_CATEGORIES)
    return {"message": "Them san pham thanh cong", "product": new_product}


@router.put("/products/{product_id}")
async def update_product(
    product_id: int,
    request: Request,
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Khong tim thay san pham")

    payload, image_file = await parse_product_payload(request)
    update_data = AdminProductUpdate(**payload)

    if update_data.name is not None:
        stripped_name = update_data.name.strip()
        # Validate tên không trùng với sản phẩm khác (case-insensitive)
        duplicate = db.query(models.Product).filter(
            func.lower(func.trim(models.Product.name)) == stripped_name.lower(),
            models.Product.id != product_id,
            models.Product.is_active == True,
        ).first()
        if duplicate:
            raise HTTPException(
                status_code=409,
                detail=f"Sản phẩm với tên '{stripped_name}' đã tồn tại. Vui lòng chọn tên khác.",
            )
        product.name = stripped_name
        if product.name:
            product.slug = slugify(product.name)
    if update_data.description is not None:
        product.description = update_data.description
    if update_data.price is not None:
        product.price = update_data.price
    if update_data.unit is not None:
        product.unit = update_data.unit or "kg"
    if update_data.rating is not None:
        product.rating = update_data.rating

    has_category_update = update_data.category_id is not None or update_data.category is not None
    if has_category_update:
        product.category_id = resolve_category_id(db, update_data.category_id, update_data.category)

    if update_data.quantity is not None:
        set_product_quantity(
            db=db,
            product=product,
            new_quantity=update_data.quantity,
            user_id=current_user.id,
            note=f"Dieu chinh so luong khi cap nhat san pham #{product.id}",
        )
    elif update_data.stock is not None:
        set_product_quantity(
            db=db,
            product=product,
            new_quantity=update_data.stock,
            user_id=current_user.id,
            note=f"Dieu chinh so luong khi cap nhat san pham #{product.id}",
        )

    if update_data.low_stock_threshold is not None:
        product.low_stock_threshold = update_data.low_stock_threshold

    next_image_url = update_data.image_url or update_data.img
    if image_file:
        next_image_url = save_uploaded_image(
            image_file,
            upload_dir=PRODUCT_UPLOAD_DIR,
            relative_dir="products",
        )
    if next_image_url:
        product.image_url = next_image_url

    product.stock_status = derive_stock_status(int(product.quantity or 0))

    # [H3] Audit log update product
    log_audit(
        db, user_id=current_user.id,
        action="updated_product", resource_type="product",
        resource_id=product_id,
        details={"name": product.name, "price": float(product.price)},
        request=request,
    )
    db.commit()
    db.refresh(product)
    # [M5] Invalidate product and stats caches
    cache_delete_pattern(CACHE_KEY_PRODUCTS_PREFIX + "*")
    cache_delete(CACHE_KEY_ADMIN_STATS)
    return {"message": "Cap nhat san pham thanh cong", "product": product}


@router.delete("/products/{product_id}")
def delete_product(
    product_id: int,
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Khong tim thay san pham")

    product.is_active = False
    log_audit(  # [H3]
        db, user_id=current_user.id,
        action="deleted_product", resource_type="product",
        resource_id=product_id,
        details={"name": product.name},
    )
    db.commit()
    # [M5] Invalidate product and stats caches
    cache_delete_pattern(CACHE_KEY_PRODUCTS_PREFIX + "*")
    cache_delete(CACHE_KEY_ADMIN_STATS)
    return {"message": "Da an san pham khoi cua hang", "product_id": product_id}


@router.get("/admin/users/paginated", response_model=schemas.AdminUserListResponse)
def get_paginated_users(
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=12, ge=1, le=100),
    search: Optional[str] = Query(default=None),
    role: Optional[str] = Query(default="all"),
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    base_query = apply_admin_user_filters(
        db.query(models.User),
        search=search,
        role=role,
    )
    total = base_query.count()
    safe_page, total_pages, has_next = build_page_meta(page, limit, total)
    items = (
        base_query.order_by(models.User.created_at.desc(), models.User.id.desc())
        .offset((safe_page - 1) * limit)
        .limit(limit)
        .all()
    )
    return {
        "items": items,
        "total": total,
        "page": safe_page,
        "limit": limit,
        "total_pages": total_pages,
        "has_next": has_next,
    }


@router.put("/admin/users/{user_id}/role", response_model=schemas.UserResponse)
def update_user_role(
    user_id: int,
    payload: schemas.AdminUserRoleUpdate,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="Nguoi dung khong ton tai")

    if target_user.id == current_user.id and payload.role != "admin":
        raise HTTPException(status_code=400, detail="Khong the tu huy quyen admin")

    role_map = {
        "admin": models.UserRole.ADMIN,
        "staff": models.UserRole.STAFF,
        "customer": models.UserRole.CUSTOMER,
    }
    old_role = target_user.role.value if hasattr(target_user.role, "value") else str(target_user.role)
    target_user.role = role_map.get(payload.role, models.UserRole.CUSTOMER)
    log_audit(  # [H3]
        db, user_id=current_user.id,
        action="updated_user_role", resource_type="user",
        resource_id=user_id,
        details={"old_role": old_role, "new_role": payload.role},
    )
    db.commit()
    db.refresh(target_user)
    return target_user


@router.get("/admin/stats")
def get_stats(
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    # [M5] Cache stats — expensive aggregation, 30 sec TTL
    cached = cache_get(CACHE_KEY_ADMIN_STATS)
    if cached is not None:
        return cached

    # [C6] Dùng SQL aggregation thay vì load tất cả vào Python
    total_revenue = db.query(func.sum(models.Order.total)).scalar() or Decimal("0")

    now = datetime.datetime.now()
    start_today = datetime.datetime.combine(now.date(), datetime.time.min)
    end_today = start_today + datetime.timedelta(days=1)

    revenue_eligible_statuses = (
        models.OrderStatus.PENDING,
        models.OrderStatus.CONFIRMED,
        models.OrderStatus.SHIPPED,
        models.OrderStatus.DELIVERED,
    )

    # [C6] Aggregate trực tiếp ở DB — không load rows
    today_revenue = (
        db.query(func.sum(models.Order.total))
        .filter(
            models.Order.created_at >= start_today,
            models.Order.created_at < end_today,
            models.Order.status.in_(revenue_eligible_statuses),
        )
        .scalar()
        or Decimal("0")
    )

    today_order_count = (
        db.query(func.count(models.Order.id))
        .filter(
            models.Order.created_at >= start_today,
            models.Order.created_at < end_today,
        )
        .scalar()
        or 0
    )

    low_stock_query = (
        db.query(models.Product)
        .options(joinedload(models.Product.category))
        .filter(
            models.Product.is_active == True,
            models.Product.quantity <= models.Product.low_stock_threshold,
        )
        .order_by(models.Product.quantity.asc(), models.Product.updated_at.desc())
    )
    low_stock_products = low_stock_query.count()

    recent_feedback_rows = (
        db.query(models.ScanFeedbackEvent)
        .options(
            joinedload(models.ScanFeedbackEvent.user),
            joinedload(models.ScanFeedbackEvent.scan),
        )
        .order_by(
            models.ScanFeedbackEvent.is_read.asc(),
            models.ScanFeedbackEvent.created_at.desc(),
        )
        .limit(5)
        .all()
    )
    unread_feedback_count = (
        db.query(models.ScanFeedbackEvent)
        .filter(get_unread_feedback_filter())
        .count()
    )

    result = {
        "total_users": db.query(models.User).count(),
        "total_products": db.query(models.Product).filter(models.Product.is_active == True).count(),
        "total_orders": db.query(models.Order).count(),
        "pending_orders": db.query(models.Order).filter(
            models.Order.status == models.OrderStatus.PENDING
        ).count(),
        "total_revenue": float(total_revenue),
        "today_revenue": float(today_revenue),
        "new_orders_today": today_order_count,
        "low_stock_products": low_stock_products,
        "unread_feedback_count": unread_feedback_count,
        "low_stock_items": [
            {
                "id": item.id,
                "name": item.name,
                "quantity": int(item.quantity or 0),
                "low_stock_threshold": int(item.low_stock_threshold or 0),
                "stock_status": item.stock_status,
                "image_url": item.image_url,
                "category_name": item.category.name if item.category else None,
            }
            for item in low_stock_query.limit(6).all()
        ],
        "recent_feedback": [serialize_feedback_event(item) for item in recent_feedback_rows],
        "system_status": "on dinh",
    }
    cache_set(CACHE_KEY_ADMIN_STATS, result, ttl=CACHE_TTL_ADMIN_STATS)
    return result


@router.get("/admin/payment-qr", response_model=Optional[schemas.PaymentQRCodeResponse])
def get_payment_qr_setting(
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    setting = (
        db.query(models.PaymentQRCodeSetting)
        .options(joinedload(models.PaymentQRCodeSetting.updated_by))
        .order_by(models.PaymentQRCodeSetting.updated_at.desc(), models.PaymentQRCodeSetting.id.desc())
        .first()
    )
    if not setting:
        return None
    return serialize_payment_qr_setting(setting)


@router.put("/admin/payment-qr", response_model=schemas.PaymentQRCodeResponse)
async def upsert_payment_qr_setting(
    request: Request,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    form_data = await request.form()
    provider_name = (form_data.get("provider_name") or "Chuyen khoan").strip() or "Chuyen khoan"
    image_file = form_data.get("image_file")

    setting = (
        db.query(models.PaymentQRCodeSetting)
        .options(joinedload(models.PaymentQRCodeSetting.updated_by))
        .order_by(models.PaymentQRCodeSetting.updated_at.desc(), models.PaymentQRCodeSetting.id.desc())
        .first()
    )

    image_url = setting.image_url if setting else None
    if image_file and isinstance(image_file, (UploadFile, StarletteUploadFile)):
        image_url = save_uploaded_image(
            image_file,
            upload_dir=PAYMENT_QR_UPLOAD_DIR,
            relative_dir="payment_qr",
        )
    elif image_url is None:
        raise HTTPException(status_code=400, detail="Vui long tai len anh ma QR hop le")

    if setting is None:
        setting = models.PaymentQRCodeSetting(
            provider_name=provider_name,
            image_url=image_url,
            updated_by_user_id=current_user.id,
        )
        db.add(setting)
    else:
        setting.provider_name = provider_name
        setting.image_url = image_url
        setting.updated_by_user_id = current_user.id

    db.commit()
    db.refresh(setting)
    return serialize_payment_qr_setting(setting)


@router.get("/admin/feedback-events", response_model=schemas.AdminScanFeedbackListResponse)
def get_admin_feedback_events(
    status: str = Query(default="all"),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=12, ge=1, le=50),
    search: Optional[str] = Query(default=None),
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    base_query = (
        db.query(models.ScanFeedbackEvent)
        .options(
            joinedload(models.ScanFeedbackEvent.user),
            joinedload(models.ScanFeedbackEvent.scan),
        )
        .outerjoin(models.User, models.ScanFeedbackEvent.user_id == models.User.id)
    )

    normalized_status = (status or "all").strip().lower()
    if normalized_status == "unread":
        base_query = base_query.filter(get_unread_feedback_filter())
    elif normalized_status == "read":
        base_query = base_query.filter(models.ScanFeedbackEvent.is_read == True)
    elif normalized_status != "all":
        raise HTTPException(status_code=400, detail="status phai la all, unread hoac read")

    if search:
        keyword = f"%{search.strip().lower()}%"
        base_query = base_query.filter(
            or_(
                func.lower(func.coalesce(models.ScanFeedbackEvent.predicted_label, "")).like(keyword),
                func.lower(func.coalesce(models.ScanFeedbackEvent.corrected_label, "")).like(keyword),
                func.lower(func.coalesce(models.ScanFeedbackEvent.notes, "")).like(keyword),
                func.lower(func.coalesce(models.ScanFeedbackEvent.source, "")).like(keyword),
                func.lower(func.coalesce(models.User.username, "")).like(keyword),
                func.lower(func.coalesce(models.User.full_name, "")).like(keyword),
                func.lower(func.coalesce(models.User.email, "")).like(keyword),
            )
        )

    unread_count = (
        db.query(models.ScanFeedbackEvent)
        .filter(get_unread_feedback_filter())
        .count()
    )
    total = base_query.count()
    safe_page, total_pages, has_next = build_page_meta(page, limit, total)
    items = (
        base_query.order_by(
            models.ScanFeedbackEvent.is_read.asc(),
            models.ScanFeedbackEvent.created_at.desc(),
        )
        .offset((safe_page - 1) * limit)
        .limit(limit)
        .all()
    )

    return {
        "items": [serialize_feedback_event(item) for item in items],
        "total": total,
        "unread_count": unread_count,
        "page": safe_page,
        "limit": limit,
        "total_pages": total_pages,
        "has_next": has_next,
    }


@router.put("/admin/feedback-events/{feedback_id}/read", response_model=schemas.AdminScanFeedbackResponse)
def mark_feedback_event_as_read(
    feedback_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    feedback_event = (
        db.query(models.ScanFeedbackEvent)
        .options(
            joinedload(models.ScanFeedbackEvent.user),
            joinedload(models.ScanFeedbackEvent.scan),
        )
        .filter(models.ScanFeedbackEvent.id == feedback_id)
        .first()
    )
    if not feedback_event:
        raise HTTPException(status_code=404, detail="Khong tim thay AI feedback")

    if not feedback_event.is_read:
        feedback_event.is_read = True
        feedback_event.read_at = datetime.datetime.utcnow()
        db.commit()
        db.refresh(feedback_event)

    return serialize_feedback_event(feedback_event)


@router.delete("/admin/users/{user_id}")
def delete_user(
    user_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Nguoi dung khong ton tai")
    if db_user.is_admin:
        raise HTTPException(status_code=400, detail="Khong the xoa tai khoan admin")

    log_audit(  # [H3]
        db, user_id=current_user.id,
        action="deleted_user", resource_type="user",
        resource_id=user_id,
        details={"username": db_user.username, "email": db_user.email},
    )
    db.delete(db_user)
    db.commit()
    return {"message": f"Da xoa nguoi dung ID {user_id}"}


@router.get("/admin/orders/paginated", response_model=schemas.AdminOrderListResponse)
def get_paginated_admin_orders(
    page: int = Query(1, ge=1),
    limit: int = Query(12, ge=1, le=100),
    search: Optional[str] = Query(default=None),
    order_status: Optional[str] = Query(default=None, alias="status"),
    payment_method: Optional[str] = Query(default=None),
    date_from: Optional[datetime.date] = Query(default=None),
    date_to: Optional[datetime.date] = Query(default=None),
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    if date_from and date_to and date_from > date_to:
        raise HTTPException(status_code=400, detail="date_from khong duoc lon hon date_to")

    filtered_query = apply_admin_order_filters(
        build_admin_orders_query(db),
        search=search,
        order_status=order_status,
        payment_method=payment_method,
        date_from=date_from,
        date_to=date_to,
    )

    total = filtered_query.count()
    total_pages = max(1, (total + limit - 1) // limit)
    safe_page = min(page, total_pages) if total else 1

    status_rows = (
        filtered_query.with_entities(models.Order.status, func.count(models.Order.id))
        .group_by(models.Order.status)
        .all()
    )
    status_counts = {status.value: 0 for status in models.OrderStatus}
    for order_status_value, count in status_rows:
        key = order_status_value.value if hasattr(order_status_value, "value") else str(order_status_value)
        status_counts[key] = count

    items = (
        filtered_query.order_by(models.Order.created_at.desc())
        .offset((safe_page - 1) * limit)
        .limit(limit)
        .all()
    )

    return {
        "items": items,
        "total": total,
        "page": safe_page,
        "limit": limit,
        "total_pages": total_pages,
        "has_next": safe_page < total_pages,
        "status_counts": status_counts,
    }


@router.get("/admin/orders/{order_id}", response_model=schemas.OrderDetail)
def get_order_detail(
    order_id: int,
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    order = (
        db.query(models.Order)
        .options(
            joinedload(models.Order.items).joinedload(models.OrderItem.product),
            joinedload(models.Order.owner),
        )
        .filter(models.Order.id == order_id)
        .first()
    )
    if not order:
        raise HTTPException(status_code=404, detail="Don hang khong ton tai")
    return order


@router.put("/admin/orders/{order_id}/status")
def update_order_status(
    order_id: int,
    status_data: schemas.OrderStatusUpdate,
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Don hang khong ton tai")

    new_status = normalize_order_status(status_data.new_status)
    old_status = order.status.value if hasattr(order.status, "value") else str(order.status)
    apply_order_status_transition(order, new_status)
    ensure_delivered_notification(db, order)
    log_audit(  # [H3]
        db, user_id=current_user.id,
        action="updated_order_status", resource_type="order",
        resource_id=order_id,
        details={"old_status": old_status, "new_status": new_status.value},
    )
    db.commit()
    db.refresh(order)
    return {"message": "Cap nhat trang thai thanh cong", "order": order}


@router.put("/admin/orders/bulk-status")
def bulk_update_order_status(
    payload: schemas.OrderBulkStatusUpdate,
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    unique_order_ids = list(dict.fromkeys(payload.order_ids))
    if not unique_order_ids:
        raise HTTPException(status_code=400, detail="Danh sach don hang khong duoc de trong")

    new_status = normalize_order_status(payload.new_status)
    orders = (
        db.query(models.Order)
        .filter(models.Order.id.in_(unique_order_ids))
        .order_by(models.Order.created_at.desc())
        .all()
    )
    if not orders:
        raise HTTPException(status_code=404, detail="Khong tim thay don hang nao de cap nhat")

    found_ids = {order.id for order in orders}
    missing_ids = [order_id for order_id in unique_order_ids if order_id not in found_ids]

    for order in orders:
        apply_order_status_transition(order, new_status)
        ensure_delivered_notification(db, order)

    db.commit()

    return {
        "message": "Cap nhat trang thai hang loat thanh cong",
        "updated_count": len(orders),
        "updated_order_ids": [order.id for order in orders],
        "missing_order_ids": missing_ids,
        "new_status": new_status.value,
    }


@router.delete("/admin/orders/{order_id}")
def delete_order(
    order_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Don hang khong ton tai")

    db.delete(order)
    db.commit()
    return {"message": f"Da xoa don hang ID {order_id}"}


@router.get("/users/{user_id}/orders", response_model=List[schemas.Order])
def get_user_orders(
    user_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.id != user_id and not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Khong co quyen xem don hang nay")
    return db.query(models.Order).filter(models.Order.user_id == user_id).all()


@router.get("/admin/orders/user/{user_id}", response_model=List[schemas.Order])
def get_user_orders_admin(
    user_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    return db.query(models.Order).filter(models.Order.user_id == user_id).all()


@router.get("/user/orders", response_model=List[schemas.Order])
def get_current_user_orders(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return db.query(models.Order).filter(models.Order.user_id == current_user.id).all()


@router.get("/user/profile", response_model=schemas.UserProfileResponse)
def get_user_profile(current_user: models.User = Depends(get_current_user)):
    return current_user


@router.put("/user/profile", response_model=schemas.UserProfileResponse)
def update_user_profile(
    user_update: schemas.UserUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    update_data = user_update.model_dump(exclude_unset=True)
    password = update_data.pop("password", None)

    for field, value in update_data.items():
        setattr(current_user, field, value)
    if password:
        current_user.hashed_password = hash_password(password)

    db.commit()
    db.refresh(current_user)
    return current_user


# ============ STOCK TRANSACTIONS ============

@router.get("/admin/stock-transactions/paginated", response_model=schemas.AdminStockTransactionListResponse)
def get_paginated_stock_transactions(
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=12, ge=1, le=100),
    search: Optional[str] = Query(default=None),
    transaction_type: Optional[str] = Query(default="all", alias="type"),
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    base_query = apply_stock_transaction_filters(
        db.query(models.StockTransaction)
        .options(
            joinedload(models.StockTransaction.product),
            joinedload(models.StockTransaction.user),
        )
        .join(models.Product, models.StockTransaction.product_id == models.Product.id)
        .join(models.User, models.StockTransaction.user_id == models.User.id),
        search=search,
        transaction_type=transaction_type,
    )
    total = base_query.count()
    safe_page, total_pages, has_next = build_page_meta(page, limit, total)
    items = (
        base_query.order_by(models.StockTransaction.created_at.desc(), models.StockTransaction.id.desc())
        .offset((safe_page - 1) * limit)
        .limit(limit)
        .all()
    )
    return {
        "items": items,
        "total": total,
        "page": safe_page,
        "limit": limit,
        "total_pages": total_pages,
        "has_next": has_next,
    }


@router.get("/admin/stock-transactions/export-excel")
def export_stock_transactions_excel(
    search: Optional[str] = Query(default=None),
    transaction_type: Optional[str] = Query(default="all", alias="type"),
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    Workbook, _ = require_openpyxl()
    transactions = (
        apply_stock_transaction_filters(
            db.query(models.StockTransaction)
            .options(
                joinedload(models.StockTransaction.product),
                joinedload(models.StockTransaction.user),
            )
            .join(models.Product, models.StockTransaction.product_id == models.Product.id)
            .join(models.User, models.StockTransaction.user_id == models.User.id),
            search=search,
            transaction_type=transaction_type,
        )
        .order_by(models.StockTransaction.created_at.desc(), models.StockTransaction.id.desc())
        .all()
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "stock_transactions"
    worksheet.append(
        [
            "transaction_id",
            "product_id",
            "product_name",
            "type",
            "quantity",
            "transaction_date",
            "created_at",
            "handled_by",
            "note",
        ]
    )

    for transaction in transactions:
        worksheet.append(
            [
                transaction.id,
                transaction.product_id,
                transaction.product.name if transaction.product else "",
                str(transaction.type),
                transaction.quantity,
                transaction.transaction_date.isoformat(sep=" ") if transaction.transaction_date else "",
                transaction.created_at.isoformat(sep=" ") if transaction.created_at else "",
                transaction.user.full_name if transaction.user and transaction.user.full_name else (
                    transaction.user.username if transaction.user else ""
                ),
                transaction.note or "",
            ]
        )

    worksheet.column_dimensions["A"].width = 16
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].width = 34
    worksheet.column_dimensions["D"].width = 14
    worksheet.column_dimensions["E"].width = 12
    worksheet.column_dimensions["F"].width = 22
    worksheet.column_dimensions["G"].width = 22
    worksheet.column_dimensions["H"].width = 24
    worksheet.column_dimensions["I"].width = 42

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"stock-transactions-{datetime.datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/admin/stock-transactions", response_model=schemas.StockTransactionResponse)
def add_stock_transaction(
    data: schemas.StockTransactionCreate,
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    product = db.query(models.Product).filter(models.Product.id == data.product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Khong tim thay san pham")

    if data.type not in ("import", "export"):
        raise HTTPException(status_code=400, detail="Loai giao dich khong hop le (import/export)")

    if data.type == "export" and (product.quantity or 0) < data.quantity:
        raise HTTPException(
            status_code=400,
            detail=f"So luong xuat ({data.quantity}) vuot qua ton kho hien tai ({product.quantity or 0})",
        )

    transaction = apply_stock_delta(
        db=db,
        product=product,
        delta=data.quantity if data.type == "import" else -data.quantity,
        user_id=current_user.id,
        note=data.note,
        transaction_date=data.transaction_date,
    )
    db.commit()
    db.refresh(product)
    if transaction is not None:
        db.refresh(transaction)
    return transaction


@router.post("/admin/stock-transactions/import-excel")
async def import_stock_transactions_excel(
    excel_file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_staff_or_admin_user),
    db: Session = Depends(get_db),
):
    if not excel_file.filename or not excel_file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Chi chap nhan file Excel .xlsx")

    _, load_workbook = require_openpyxl()

    try:
        file_bytes = await excel_file.read()
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        worksheet = workbook.active
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Khong the doc file Excel. Vui long kiem tra lai dinh dang.") from exc

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File Excel khong co du lieu.")

    header_aliases = {
        "product_id": "product_id",
        "ma_san_pham": "product_id",
        "id_san_pham": "product_id",
        "type": "type",
        "transaction_type": "type",
        "loai": "type",
        "quantity": "quantity",
        "so_luong": "quantity",
        "transaction_date": "transaction_date",
        "ngay_giao_dich": "transaction_date",
        "date": "transaction_date",
        "note": "note",
        "ghi_chu": "note",
    }

    header_row = rows[0]
    column_map = {}
    for index, header in enumerate(header_row):
        alias = header_aliases.get(normalize_excel_header(header))
        if alias and alias not in column_map:
            column_map[alias] = index

    missing_headers = [header for header in ("product_id", "type", "quantity") if header not in column_map]
    if missing_headers:
        raise HTTPException(
            status_code=400,
            detail=f"File Excel thieu cot bat buoc: {', '.join(missing_headers)}",
        )

    operations = []
    parse_errors: List[str] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not row or all(cell in (None, "") for cell in row):
            continue

        try:
            product_id_raw = row[column_map["product_id"]]
            quantity_raw = row[column_map["quantity"]]
            transaction_type_raw = row[column_map["type"]]
            transaction_date_raw = row[column_map["transaction_date"]] if "transaction_date" in column_map else None
            note_raw = row[column_map["note"]] if "note" in column_map else None

            product_id = int(product_id_raw)
            quantity = int(quantity_raw)
            if quantity < 1:
                raise ValueError("So luong phai lon hon 0")

            operations.append(
                {
                    "row_number": row_number,
                    "product_id": product_id,
                    "type": normalize_excel_transaction_type(transaction_type_raw),
                    "quantity": quantity,
                    "transaction_date": parse_excel_transaction_date(transaction_date_raw),
                    "note": str(note_raw).strip() if note_raw not in (None, "") else None,
                }
            )
        except ValueError as exc:
            parse_errors.append(f"Dong {row_number}: {exc}")

    if parse_errors:
        raise HTTPException(
            status_code=400,
            detail=build_excel_error_detail("File Excel co du lieu khong hop le:", parse_errors),
        )

    if not operations:
        raise HTTPException(status_code=400, detail="File Excel khong co dong giao dich hop le nao.")

    product_ids = sorted({item["product_id"] for item in operations})
    products = (
        db.query(models.Product)
        .filter(models.Product.id.in_(product_ids))
        .all()
    )
    products_by_id = {product.id: product for product in products}

    missing_products = [product_id for product_id in product_ids if product_id not in products_by_id]
    if missing_products:
        raise HTTPException(
            status_code=400,
            detail=f"Khong tim thay san pham voi ID: {', '.join(str(item) for item in missing_products[:12])}",
        )

    validation_errors: List[str] = []
    stock_projection = {product.id: int(product.quantity or 0) for product in products}
    for item in operations:
        current_quantity = stock_projection[item["product_id"]]
        delta = item["quantity"] if item["type"] == "import" else -item["quantity"]
        next_quantity = current_quantity + delta
        if next_quantity < 0:
            validation_errors.append(
                f"Dong {item['row_number']}: xuat {item['quantity']} vuot ton kho hien tai ({current_quantity}) cua san pham #{item['product_id']}"
            )
            continue
        stock_projection[item["product_id"]] = next_quantity

    if validation_errors:
        raise HTTPException(
            status_code=400,
            detail=build_excel_error_detail("Khong the nhap file Excel vi ton kho khong hop le:", validation_errors),
        )

    imported_count = 0
    try:
        for item in operations:
            product = products_by_id[item["product_id"]]
            delta = item["quantity"] if item["type"] == "import" else -item["quantity"]
            apply_stock_delta(
                db=db,
                product=product,
                delta=delta,
                user_id=current_user.id,
                note=item["note"] or f"Import Excel dong {item['row_number']}",
                transaction_date=item["transaction_date"],
            )
            imported_count += 1

        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail="Khong the nhap giao dich tu file Excel.") from exc

    return {
        "imported_count": imported_count,
        "product_count": len(product_ids),
        "message": "Nhap file Excel thanh cong",
    }


# ============ FRESHNESS COMPLAINTS — ADMIN APPROVE/REJECT ============

@router.get("/admin/freshness-complaints")
def list_freshness_complaints(
    resolution_status: Optional[str] = Query(default="pending_review"),
    page: int = Query(default=1, ge=1),
    limit: int = Query(default=12, ge=1, le=100),
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    """[C1] Danh sách khiếu nại độ tươi — admin xem để phê duyệt refund."""
    from sqlalchemy.orm import joinedload as jl
    query = (
        db.query(models.FreshnessComplaint)
        .options(
            jl(models.FreshnessComplaint.user),
            jl(models.FreshnessComplaint.order),
        )
    )
    if resolution_status and resolution_status != "all":
        query = query.filter(models.FreshnessComplaint.resolution_status == resolution_status)

    total = query.count()
    items = query.order_by(models.FreshnessComplaint.created_at.desc()).offset((page - 1) * limit).limit(limit).all()
    return {"items": items, "total": total, "page": page, "limit": limit}


@router.put("/admin/freshness-complaints/{complaint_id}/approve")
def approve_freshness_complaint(
    complaint_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    """[C1] Admin phê duyệt refund — thực sự cộng voucher vào tài khoản user."""
    complaint = db.query(models.FreshnessComplaint).filter(
        models.FreshnessComplaint.id == complaint_id
    ).first()
    if not complaint:
        raise HTTPException(status_code=404, detail="Khong tim thay khieu nai")
    if complaint.resolution_status != ComplaintResolutionStatus.PENDING_REVIEW:
        raise HTTPException(status_code=400, detail=f"Khieu nai khong o trang thai cho duyet (hien tai: {complaint.resolution_status})")

    user = db.query(models.User).filter(models.User.id == complaint.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Nguoi dung khong ton tai")

    if complaint.complaint_type == "refund" and complaint.refund_amount:
        user.voucher_balance = Decimal(user.voucher_balance or 0) + complaint.refund_amount

    complaint.resolution_status = ComplaintResolutionStatus.APPROVED
    db.commit()
    return {"success": True, "message": "Da phe duyet va cong voucher thanh cong", "complaint_id": complaint_id}


@router.put("/admin/freshness-complaints/{complaint_id}/reject")
def reject_freshness_complaint(
    complaint_id: int,
    current_user: models.User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    """[C1] Admin từ chối refund."""
    complaint = db.query(models.FreshnessComplaint).filter(
        models.FreshnessComplaint.id == complaint_id
    ).first()
    if not complaint:
        raise HTTPException(status_code=404, detail="Khong tim thay khieu nai")
    if complaint.resolution_status != ComplaintResolutionStatus.PENDING_REVIEW:
        raise HTTPException(status_code=400, detail="Khieu nai khong o trang thai cho duyet")

    complaint.resolution_status = ComplaintResolutionStatus.REJECTED
    db.commit()
    return {"success": True, "message": "Da tu choi khieu nai", "complaint_id": complaint_id}
